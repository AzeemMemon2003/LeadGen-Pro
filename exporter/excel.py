from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:

    def __init__(self):

        self.workbook = Workbook()

        self.sheet = self.workbook.active

        self.sheet.title = "LeadGen Pro"

        headers = [
            "Company",
            "Website",
            "Title",
            "Emails",
            "Phones",
            "Addresses",
            "Technology",
            "LinkedIn",
            "Facebook",
            "Instagram",
            "Twitter",
            "Website Score",
            "Strengths",
            "Weaknesses",
            "Sales Opportunities",
            "Contact Pages"
        ]

        self.sheet.append(headers)

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for cell in self.sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        self.sheet.freeze_panes = "A2"

        self.sheet.auto_filter.ref = "A1:P1"

    def add(
        self,
        company,
        website,
        title,
        emails,
        phones,
        addresses,
        technology,
        social,
        contacts,
        website_intelligence
    ):

        self.sheet.append([

            company,

            website,

            title,

            ", ".join(emails),

            ", ".join(phones),

            ", ".join(addresses),

            ", ".join(technology),

            social.get("linkedin", ""),

            social.get("facebook", ""),

            social.get("instagram", ""),

            social.get("twitter", ""),

            website_intelligence.get("website_score", 0),

            ", ".join(
                website_intelligence.get("strengths", [])
            ),

            ", ".join(
                website_intelligence.get("weaknesses", [])
            ),

            ", ".join(
                website_intelligence.get("sales_opportunities", [])
            ),

            ", ".join(contacts)

        ])

    def save(self):

        for column in self.sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            self.sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 3, 60)

        from config.settings import Settings